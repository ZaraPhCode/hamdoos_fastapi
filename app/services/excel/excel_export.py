"""Excel export service — products, orders, invoices.

Mirrors the Excel export functionality from Invoice.cs (ConvertXLSX).
Uses openpyxl.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _style_header(ws, headers: list[str], row: int = 1):
    """Style the header row."""
    fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def export_products_to_excel(products: list[dict]) -> io.BytesIO:
    """Export products list to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["Name", "Part Number", "Model", "Category", "Brand", "Price", "Stock", "Status"]
    _style_header(ws, headers)

    for i, p in enumerate(products, 2):
        ws.cell(row=i, column=1, value=p.get("name", ""))
        ws.cell(row=i, column=2, value=p.get("part_number", ""))
        ws.cell(row=i, column=3, value=p.get("model", ""))
        ws.cell(row=i, column=4, value=p.get("category_title", ""))
        ws.cell(row=i, column=5, value=p.get("brand_name", ""))
        ws.cell(row=i, column=6, value=p.get("price", 0))
        ws.cell(row=i, column=7, value=p.get("stock_quantity", 0))
        ws.cell(row=i, column=8, value=p.get("status", ""))

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_orders_to_excel(orders: list[dict]) -> io.BytesIO:
    """Export orders list to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["Reference Code", "Customer", "Status", "Total", "Payable", "Date"]
    _style_header(ws, headers)

    for i, o in enumerate(orders, 2):
        ws.cell(row=i, column=1, value=o.get("reference_code", ""))
        ws.cell(row=i, column=2, value=f"{o.get('first_name', '')} {o.get('last_name', '')}")
        ws.cell(row=i, column=3, value=o.get("order_status", ""))
        ws.cell(row=i, column=4, value=o.get("total_price", 0))
        ws.cell(row=i, column=5, value=o.get("payable", 0))
        ws.cell(row=i, column=6, value=str(o.get("insert_date", "")))

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_invoices_to_excel(invoices: list[dict]) -> io.BytesIO:
    """Export invoices list to Excel. Mirrors Invoice.ConvertXLSX()."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = ["Ref Code", "Type", "Customer", "Total", "Tax", "Payable", "Status", "Date"]
    _style_header(ws, headers)

    for i, inv in enumerate(invoices, 2):
        ws.cell(row=i, column=1, value=inv.get("reference_code", ""))
        ws.cell(row=i, column=2, value=inv.get("type", ""))
        ws.cell(row=i, column=3, value=inv.get("identity_name", ""))
        ws.cell(row=i, column=4, value=inv.get("total_price", 0))
        ws.cell(row=i, column=5, value=inv.get("total_taxes_and_duties", 0))
        ws.cell(row=i, column=6, value=inv.get("payable", 0))
        ws.cell(row=i, column=7, value=inv.get("status", ""))
        ws.cell(row=i, column=8, value=str(inv.get("date", "")))

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf